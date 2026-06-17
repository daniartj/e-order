import base64
import calendar
import datetime
import json

from django.conf import settings
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import login_required
from django.db.models import Sum
from django.http import HttpResponse, JsonResponse
from django.shortcuts import redirect, render
from django.utils import timezone
from django.views.decorators.http import require_POST

from menu.models import (DetailPesanan, KategoriMenu, Meja, Menu, Pengaturan,
                         Pesanan)
from menu.utils_qr import generate_qr_dengan_label
from pelanggan.utils import get_client_ip
from menu.models import PushSubscription
from django.conf import settings


# ==============================
# LOGIN KASIR
# ==============================
def login_kasir(request):
    if request.user.is_authenticated:
        return redirect('kasir:dashboard')

    error = None

    if request.method == 'POST':
        username = request.POST.get('username', '').strip()
        password = request.POST.get('password', '').strip()

        if not username or not password:
            error = 'Username dan password wajib diisi.'
        else:
            user = authenticate(request, username=username, password=password)

            if user is not None:
                login(request, user)
                next_url = request.GET.get('next', 'kasir:dashboard')
                return redirect(next_url)
            else:
                error = 'Username atau password salah.'

    return render(request, 'kasir/login.html', {
        'error': error
    })


# ==============================
# LOGOUT KASIR
# ==============================
def logout_kasir(request):
    logout(request)
    return redirect('kasir:login')


# ==============================
# DASHBOARD KASIR
# ==============================
@login_required(login_url='/kasir/login/')
def dashboard(request):
    filter_status = request.GET.get('status', 'semua')

    if filter_status == 'semua':
        daftar_pesanan = Pesanan.objects.prefetch_related(
            'detail'
        ).select_related('meja').order_by('-waktu_pesan')
    else:
        daftar_pesanan = Pesanan.objects.prefetch_related(
            'detail'
        ).select_related('meja').filter(
            status=filter_status
        ).order_by('-waktu_pesan')

    jumlah_pending  = Pesanan.objects.filter(status='pending').count()
    jumlah_diproses = Pesanan.objects.filter(status='diproses').count()
    jumlah_selesai  = Pesanan.objects.filter(status='selesai').count()

    filter_tabs = [
        {'value': 'semua',    'label': '📋 Semua'},
        {'value': 'pending',  'label': '⏳ Pending'},
        {'value': 'diproses', 'label': '👨‍🍳 Diproses'},
        {'value': 'selesai',  'label': '✅ Selesai'},
        {'value': 'dibatal',  'label': '❌ Dibatal'},
    ]

    # ✅ Kirim setting ke template
    setting = Pengaturan.get()

    return render(request, 'kasir/dashboard.html', {
        'daftar_pesanan' : daftar_pesanan,
        'filter_status'  : filter_status,
        'jumlah_pending' : jumlah_pending,
        'jumlah_diproses': jumlah_diproses,
        'jumlah_selesai' : jumlah_selesai,
        'filter_tabs'    : filter_tabs,
        'setting'        : setting,
    })


# ==============================
# UPDATE STATUS PESANAN
# ==============================
@login_required(login_url='/kasir/login/')
@require_POST
def update_status(request, nomor_pesanan):
    """
    Update status pesanan + kirim notifikasi WhatsApp
    jika pelanggan menyediakan nomor WA.
    """
    from menu.whatsapp import kirim_notifikasi, pesan_diproses, pesan_selesai

    try:
        data        = json.loads(request.body)
        status_baru = data.get('status')

        STATUS_VALID = ['pending', 'diproses', 'selesai', 'dibatal']
        if status_baru not in STATUS_VALID:
            return JsonResponse({
                'status': 'error',
                'pesan' : 'Status tidak valid.'
            }, status=400)

        pesanan        = Pesanan.objects.get(nomor_pesanan=nomor_pesanan)
        pesanan.status = status_baru

        if status_baru == 'selesai':
            pesanan.waktu_selesai = timezone.now()

        pesanan.save()

        # ==============================
        # PUSH NOTIFIKASI (PENDING BARU)
        # ==============================
        if status_baru == 'pending':
            try:
                import logging
                logger = logging.getLogger(__name__)
                logger.warning(
                    "[PUSH pending] trigger kirim_push_semua_kasir. nomor_pesanan=%s meja=%s user=%s",
                    pesanan.nomor_pesanan,
                    getattr(pesanan.meja, 'nomor_meja', None),
                    getattr(getattr(request, 'user', None), 'username', None),
                )

                from kasir.push import kirim_push_semua_kasir
                kirim_push_semua_kasir(
                    judul='🔔 Pesanan Baru',
                    isi=f'Pesanan {pesanan.nomor_pesanan} masuk.',
                    url=f'/kasir/dashboard/?status=pending'
                )
            except Exception as e:
                import logging
                logger = logging.getLogger(__name__)
                logger.exception(
                    "[PUSH pending] gagal kirim push. nomor_pesanan=%s meja=%s user=%s err=%s",
                    pesanan.nomor_pesanan,
                    getattr(pesanan.meja, 'nomor_meja', None),
                    getattr(getattr(request, 'user', None), 'username', None),
                    str(e),
                )



        # ==============================
        # KIRIM NOTIFIKASI WHATSAPP
        # ==============================
        hasil_wa  = None
        nomor_meja = pesanan.meja.nomor_meja if pesanan.meja else '-'

        if (pesanan.nomor_whatsapp and
                status_baru in ['diproses', 'selesai'] and
                getattr(settings, 'WHATSAPP_AKTIF', False)):

            if status_baru == 'diproses':
                isi_pesan = pesan_diproses(nomor_pesanan, nomor_meja)
            else:
                isi_pesan = pesan_selesai(nomor_pesanan, nomor_meja)

            hasil_wa = kirim_notifikasi(
                pesanan.nomor_whatsapp,
                isi_pesan
            )

        return JsonResponse({
            'status'      : 'ok',
            'pesan'       : f'Status diubah ke {pesanan.get_status_display()}',
            'status_baru' : pesanan.status,
            'label_status': pesanan.get_status_display(),
            'wa_terkirim' : hasil_wa.get('sukses', False) if hasil_wa else False,
            'wa_pesan'    : hasil_wa.get('pesan', '') if hasil_wa else '',
        })

    except Pesanan.DoesNotExist:
        return JsonResponse({
            'status': 'error',
            'pesan' : 'Pesanan tidak ditemukan.'
        }, status=404)

    except Exception as e:
        return JsonResponse({
            'status': 'error',
            'pesan' : str(e)
        }, status=500)


# ==============================
# DETAIL PESANAN (MODAL)
# ==============================
@login_required(login_url='/kasir/login/')
def detail_pesanan(request, nomor_pesanan):
    try:
        pesanan = Pesanan.objects.prefetch_related(
            'detail'
        ).select_related('meja').get(nomor_pesanan=nomor_pesanan)

        detail_list = []
        for item in pesanan.detail.all():
            detail_list.append({
                'nama_menu': item.nama_menu_snapshot,
                'jumlah'   : item.jumlah,
                'harga'    : item.harga_snapshot,
                'subtotal' : item.subtotal,
            })

        return JsonResponse({
            'status'        : 'ok',
            'nomor_pesanan' : pesanan.nomor_pesanan,
            'nomor_meja'    : pesanan.meja.nomor_meja if pesanan.meja else '-',
            'nama_meja'     : pesanan.meja.nama_meja if pesanan.meja else '-',
            'status_pesanan': pesanan.status,
            'label_status'  : pesanan.get_status_display(),
            'metode_bayar'  : pesanan.get_metode_bayar_display(),
            'total_harga'   : pesanan.total_harga,
            'total_format'  : pesanan.total_format(),
            'catatan'       : pesanan.catatan,
            'nomor_wa'      : pesanan.nomor_whatsapp or '-',
            'waktu_pesan'   : pesanan.waktu_pesan.strftime('%H:%M WIB'),
            'detail'        : detail_list,
        })

    except Pesanan.DoesNotExist:
        return JsonResponse({
            'status': 'error',
            'pesan' : 'Pesanan tidak ditemukan.'
        }, status=404)


# ==============================
# ✅ API REALTIME DASHBOARD
# ==============================
@login_required(login_url='/kasir/login/')
def api_daftar_pesanan(request):
    """
    API endpoint untuk polling realtime dashboard kasir.
    Dipanggil Alpine.js setiap 5 detik.
    """
    filter_status = request.GET.get('status', 'semua')

    if filter_status == 'semua':
        daftar = Pesanan.objects.prefetch_related(
            'detail'
        ).select_related('meja').order_by('-waktu_pesan')
    else:
        daftar = Pesanan.objects.prefetch_related(
            'detail'
        ).select_related('meja').filter(
            status=filter_status
        ).order_by('-waktu_pesan')

    hasil = []
    for p in daftar:
        item_list = []
        for item in p.detail.all():
            item_list.append({
                'nama_menu': item.nama_menu_snapshot,
                'jumlah'   : item.jumlah,
            })

        hasil.append({
            'nomor_pesanan' : p.nomor_pesanan,
            'nomor_meja'    : p.meja.nomor_meja if p.meja else '-',
            'nama_meja'     : p.meja.nama_meja if p.meja else '-',
            'status'        : p.status,
            'label_status'  : p.get_status_display(),
            'metode_bayar'  : p.get_metode_bayar_display(),
            'total_format'  : p.total_format(),
            'total_harga'   : p.total_harga,
            'waktu_pesan'   : p.waktu_pesan.strftime('%H:%M'),
            'catatan'       : p.catatan,
            'nomor_whatsapp': p.nomor_whatsapp or '',
            'detail'        : item_list,
        })

    return JsonResponse({
        'status'         : 'ok',
        'daftar_pesanan' : hasil,
        'jumlah_pending' : Pesanan.objects.filter(status='pending').count(),
        'jumlah_diproses': Pesanan.objects.filter(status='diproses').count(),
        'jumlah_selesai' : Pesanan.objects.filter(status='selesai').count(),
    })

###---NOTIF PWA---######

@login_required(login_url='/kasir/login/')
@require_POST
def push_subscribe(request):
    """Simpan subscription push notification."""
    try:
        data     = json.loads(request.body)
        endpoint = data.get('endpoint', '')
        p256dh   = data.get('keys', {}).get('p256dh', '')
        auth     = data.get('keys', {}).get('auth', '')

        if not endpoint or not p256dh or not auth:
            return JsonResponse({
                'status': 'error',
                'pesan' : 'Data subscription tidak lengkap'
            }, status=400)

        # Upsert subscription
        PushSubscription.objects.update_or_create(
            endpoint = endpoint,
            defaults = {
                'user'  : request.user,
                'p256dh': p256dh,
                'auth'  : auth,
            }
        )

        return JsonResponse({'status': 'ok'})

    except Exception as e:
        return JsonResponse({
            'status': 'error',
            'pesan' : str(e)
        }, status=500)


@login_required(login_url='/kasir/login/')
@require_POST
def push_unsubscribe(request):
    """Hapus subscription push notification."""
    try:
        data     = json.loads(request.body)
        endpoint = data.get('endpoint', '')
        PushSubscription.objects.filter(endpoint=endpoint).delete()
        return JsonResponse({'status': 'ok'})
    except Exception as e:
        return JsonResponse({
            'status': 'error',
            'pesan' : str(e)
        }, status=500)


@login_required(login_url='/kasir/login/')
def push_vapid_public_key(request):
    """Return VAPID public key untuk client."""
    return JsonResponse({
        'vapid_public_key': settings.VAPID_PUBLIC_KEY
    })
# ==============================
# KELOLA MEJA
# ==============================
@login_required(login_url='/kasir/login/')
def kelola_meja(request):
    semua_meja = Meja.objects.all().order_by('nomor_meja')
    return render(request, 'kasir/kelola_meja.html', {
        'semua_meja': semua_meja,
    })


@login_required(login_url='/kasir/login/')
def tambah_meja(request):
    error = None
    if request.method == 'POST':
        nomor_meja = request.POST.get('nomor_meja', '').strip()
        nama_meja  = request.POST.get('nama_meja', '').strip()
        kapasitas  = request.POST.get('kapasitas', '4').strip()

        if not nomor_meja or not nama_meja:
            error = 'Nomor meja dan nama meja wajib diisi.'
        elif Meja.objects.filter(nomor_meja=nomor_meja).exists():
            error = f'Nomor meja {nomor_meja} sudah ada.'
        else:
            Meja.objects.create(
                nomor_meja = nomor_meja,
                nama_meja  = nama_meja,
                kapasitas  = int(kapasitas) if kapasitas.isdigit() else 4,
                is_aktif   = True,
            )
            return redirect('kasir:kelola_meja')

    return render(request, 'kasir/form_meja.html', {
        'error' : error,
        'aksi'  : 'Tambah',
        'meja'  : None,
    })


@login_required(login_url='/kasir/login/')
def edit_meja(request, meja_id):
    try:
        meja = Meja.objects.get(id=meja_id)
    except Meja.DoesNotExist:
        return redirect('kasir:kelola_meja')

    error = None
    if request.method == 'POST':
        nama_meja = request.POST.get('nama_meja', '').strip()
        kapasitas = request.POST.get('kapasitas', '4').strip()

        if not nama_meja:
            error = 'Nama meja wajib diisi.'
        else:
            meja.nama_meja = nama_meja
            meja.kapasitas = int(kapasitas) if kapasitas.isdigit() else 4
            meja.save()
            return redirect('kasir:kelola_meja')

    return render(request, 'kasir/form_meja.html', {
        'error': error,
        'aksi' : 'Edit',
        'meja' : meja,
    })


@login_required(login_url='/kasir/login/')
def toggle_meja(request, meja_id):
    try:
        meja          = Meja.objects.get(id=meja_id)
        meja.is_aktif = not meja.is_aktif
        meja.save()
    except Meja.DoesNotExist:
        pass
    return redirect('kasir:kelola_meja')


# ==============================
# KELOLA KATEGORI
# ==============================
@login_required(login_url='/kasir/login/')
def kelola_kategori(request):
    semua_kategori = KategoriMenu.objects.all().order_by('urutan')
    return render(request, 'kasir/kelola_kategori.html', {
        'semua_kategori': semua_kategori,
    })


@login_required(login_url='/kasir/login/')
def tambah_kategori(request):
    error = None
    if request.method == 'POST':
        nama = request.POST.get('nama_kategori', '').strip()
        ikon = request.POST.get('ikon', '🍽️').strip()
        urutan = request.POST.get('urutan', '0').strip()

        if not nama:
            error = 'Nama kategori wajib diisi.'
        elif KategoriMenu.objects.filter(nama_kategori=nama).exists():
            error = f'Kategori "{nama}" sudah ada.'
        else:
            KategoriMenu.objects.create(
                nama_kategori = nama,
                ikon          = ikon or '🍽️',
                urutan        = int(urutan) if urutan.isdigit() else 0,
                is_aktif      = True,
            )
            return redirect('kasir:kelola_kategori')

    return render(request, 'kasir/form_kategori.html', {
        'error'   : error,
        'aksi'    : 'Tambah',
        'kategori': None,
    })


@login_required(login_url='/kasir/login/')
def edit_kategori(request, kategori_id):
    try:
        kategori = KategoriMenu.objects.get(id=kategori_id)
    except KategoriMenu.DoesNotExist:
        return redirect('kasir:kelola_kategori')

    error = None
    if request.method == 'POST':
        nama   = request.POST.get('nama_kategori', '').strip()
        ikon   = request.POST.get('ikon', '🍽️').strip()
        urutan = request.POST.get('urutan', '0').strip()

        if not nama:
            error = 'Nama kategori wajib diisi.'
        else:
            kategori.nama_kategori = nama
            kategori.ikon          = ikon or '🍽️'
            kategori.urutan        = int(urutan) if urutan.isdigit() else 0
            kategori.save()
            return redirect('kasir:kelola_kategori')

    return render(request, 'kasir/form_kategori.html', {
        'error'   : error,
        'aksi'    : 'Edit',
        'kategori': kategori,
    })


@login_required(login_url='/kasir/login/')
def toggle_kategori(request, kategori_id):
    try:
        kat          = KategoriMenu.objects.get(id=kategori_id)
        kat.is_aktif = not kat.is_aktif
        kat.save()
    except KategoriMenu.DoesNotExist:
        pass
    return redirect('kasir:kelola_kategori')


# ==============================
# KELOLA MENU
# ==============================
@login_required(login_url='/kasir/login/')
def kelola_menu(request):
    filter_kategori = request.GET.get('kategori', '')
    semua_menu      = Menu.objects.select_related(
        'kategori'
    ).order_by('kategori', 'nama_menu')

    if filter_kategori:
        semua_menu = semua_menu.filter(kategori_id=filter_kategori)

    semua_kategori = KategoriMenu.objects.filter(is_aktif=True)

    return render(request, 'kasir/kelola_menu.html', {
        'semua_menu'     : semua_menu,
        'semua_kategori' : semua_kategori,
        'filter_kategori': filter_kategori,
    })


@login_required(login_url='/kasir/login/')
def tambah_menu(request):
    semua_kategori = KategoriMenu.objects.filter(is_aktif=True)
    error          = None

    if request.method == 'POST':
        nama       = request.POST.get('nama_menu', '').strip()
        deskripsi  = request.POST.get('deskripsi', '').strip()
        harga      = request.POST.get('harga', '').strip()
        kategori_id= request.POST.get('kategori', '')
        is_unggulan= request.POST.get('is_unggulan') == 'on'
        foto       = request.FILES.get('foto')

        if not nama or not harga:
            error = 'Nama menu dan harga wajib diisi.'
        elif not harga.isdigit():
            error = 'Harga harus berupa angka.'
        else:
            menu = Menu(
                nama_menu   = nama,
                deskripsi   = deskripsi,
                harga       = int(harga),
                is_unggulan = is_unggulan,
                is_tersedia = True,
            )
            if kategori_id:
                try:
                    menu.kategori = KategoriMenu.objects.get(
                        id=kategori_id
                    )
                except KategoriMenu.DoesNotExist:
                    pass
            if foto:
                menu.foto = foto
            menu.save()
            return redirect('kasir:kelola_menu')

    return render(request, 'kasir/form_menu.html', {
        'error'         : error,
        'aksi'          : 'Tambah',
        'menu'          : None,
        'semua_kategori': semua_kategori,
    })


@login_required(login_url='/kasir/login/')
def edit_menu(request, menu_id):
    try:
        menu = Menu.objects.get(id=menu_id)
    except Menu.DoesNotExist:
        return redirect('kasir:kelola_menu')

    semua_kategori = KategoriMenu.objects.filter(is_aktif=True)
    error          = None

    if request.method == 'POST':
        nama        = request.POST.get('nama_menu', '').strip()
        deskripsi   = request.POST.get('deskripsi', '').strip()
        harga       = request.POST.get('harga', '').strip()
        kategori_id = request.POST.get('kategori', '')
        is_unggulan = request.POST.get('is_unggulan') == 'on'
        foto        = request.FILES.get('foto')

        if not nama or not harga:
            error = 'Nama menu dan harga wajib diisi.'
        elif not harga.isdigit():
            error = 'Harga harus berupa angka.'
        else:
            menu.nama_menu   = nama
            menu.deskripsi   = deskripsi
            menu.harga       = int(harga)
            menu.is_unggulan = is_unggulan
            if kategori_id:
                try:
                    menu.kategori = KategoriMenu.objects.get(
                        id=kategori_id
                    )
                except KategoriMenu.DoesNotExist:
                    pass
            if foto:
                menu.foto = foto
            menu.save()
            return redirect('kasir:kelola_menu')

    return render(request, 'kasir/form_menu.html', {
        'error'         : error,
        'aksi'          : 'Edit',
        'menu'          : menu,
        'semua_kategori': semua_kategori,
    })


@login_required(login_url='/kasir/login/')
def toggle_menu(request, menu_id):
    try:
        menu              = Menu.objects.get(id=menu_id)
        menu.is_tersedia  = not menu.is_tersedia
        menu.save()
    except Menu.DoesNotExist:
        pass
    return redirect('kasir:kelola_menu')


NAMA_BULAN = {
    1:'Januari', 2:'Februari', 3:'Maret',    4:'April',
    5:'Mei',     6:'Juni',     7:'Juli',      8:'Agustus',
    9:'September',10:'Oktober',11:'November', 12:'Desember'
}

def _fmt_tgl(tgl):
    return tgl.strftime(f'%d {NAMA_BULAN[tgl.month]} %Y')


# ==============================
# LAPORAN TRANSAKSI — HARI INI
# ==============================
@login_required(login_url='/kasir/login/')
def lt_hari_ini(request):
    today    = timezone.localtime(timezone.now()).date()
    pesanan  = Pesanan.objects.filter(
        waktu_pesan__date=today
    ).select_related('meja').prefetch_related('detail')

    labels, val_jml, val_pend = [], [], []
    for jam in range(24):
        p = pesanan.filter(waktu_pesan__hour=jam)
        jml  = p.count()
        pend = p.filter(status='selesai').aggregate(
            t=Sum('total_harga'))['t'] or 0
        labels.append(f'{jam:02d}:00')
        val_jml.append(jml)
        val_pend.append(pend)

    return render(request, 'kasir/laporan_transaksi_hari_ini.html', {
        'pesanan'  : pesanan.order_by('-waktu_pesan'),
        'labels'   : json.dumps(labels),
        'val_jml'  : json.dumps(val_jml),
        'val_pend' : json.dumps(val_pend),
        'total_transaksi' : pesanan.count(),
        'total_selesai'   : pesanan.filter(status='selesai').count(),
        'total_pendapatan': pesanan.filter(status='selesai').aggregate(
            t=Sum('total_harga'))['t'] or 0,
        'judul' : f'Hari Ini — {_fmt_tgl(today)}',
    })


# ==============================
# LAPORAN TRANSAKSI — BULAN INI
# ==============================
@login_required(login_url='/kasir/login/')
def lt_bulan_ini(request):
    now  = timezone.localtime(timezone.now())
    ftgl = request.GET.get('tanggal', '')

    pesanan_bulan = Pesanan.objects.filter(
        waktu_pesan__year=now.year,
        waktu_pesan__month=now.month,
    ).select_related('meja').prefetch_related('detail')

    # Drill-down ke tanggal
    pesanan_tgl, tgl_dipilih = None, None
    if ftgl:
        try:
            tgl_dipilih = datetime.date.fromisoformat(ftgl)
            pesanan_tgl = pesanan_bulan.filter(
                waktu_pesan__date=tgl_dipilih
            ).order_by('-waktu_pesan')
        except ValueError:
            pass

    days = calendar.monthrange(now.year, now.month)[1]
    labels, val_jml, val_pend, data_tgl = [], [], [], []
    for day in range(1, days + 1):
        p    = pesanan_bulan.filter(waktu_pesan__day=day)
        jml  = p.count()
        pend = p.filter(status='selesai').aggregate(
            t=Sum('total_harga'))['t'] or 0
        labels.append(str(day))
        val_jml.append(jml)
        val_pend.append(pend)
        data_tgl.append({
            'hari': day,
            'tgl_str': f'{now.year}-{now.month:02d}-{day:02d}',
            'jumlah': jml, 'pendapatan': pend,
        })

    return render(request, 'kasir/laporan_transaksi_bulan_ini.html', {
        'data_tgl'        : data_tgl,
        'pesanan_tgl'     : pesanan_tgl,
        'tgl_dipilih'     : tgl_dipilih,
        'labels'          : json.dumps(labels),
        'val_jml'         : json.dumps(val_jml),
        'val_pend'        : json.dumps(val_pend),
        'total_transaksi' : pesanan_bulan.count(),
        'total_pendapatan': pesanan_bulan.filter(status='selesai').aggregate(
            t=Sum('total_harga'))['t'] or 0,
        'judul': f'{NAMA_BULAN[now.month]} {now.year}',
    })


# ==============================
# LAPORAN TRANSAKSI — TAHUN INI
# ==============================
@login_required(login_url='/kasir/login/')
def lt_tahun_ini(request):
    now   = timezone.localtime(timezone.now())
    fbulan = request.GET.get('bulan', '')
    ftgl   = request.GET.get('tanggal', '')

    pesanan_tahun = Pesanan.objects.filter(
        waktu_pesan__year=now.year
    ).select_related('meja').prefetch_related('detail')

    # Chart per bulan
    labels_b, val_jml_b, val_pend_b, data_bulan = [], [], [], []
    for bln in range(1, 13):
        p    = pesanan_tahun.filter(waktu_pesan__month=bln)
        jml  = p.count()
        pend = p.filter(status='selesai').aggregate(
            t=Sum('total_harga'))['t'] or 0
        data_bulan.append({
            'bulan': bln, 'nama': NAMA_BULAN[bln],
            'bln_str': f'{now.year}-{bln:02d}',
            'jumlah': jml, 'pendapatan': pend,
        })
        labels_b.append(NAMA_BULAN[bln][:3])
        val_jml_b.append(jml)
        val_pend_b.append(pend)

    # Drill-down ke tanggal dalam bulan
    data_tgl_bln, bln_dipilih = None, None
    labels_t, val_jml_t, val_pend_t = [], [], []

    if fbulan and not ftgl:
        try:
            thn_b, bln_b = int(fbulan.split('-')[0]), int(fbulan.split('-')[1])
            bln_dipilih  = {'str': fbulan, 'nama': NAMA_BULAN[bln_b], 'tahun': thn_b, 'bulan': bln_b}
            p_bln = pesanan_tahun.filter(waktu_pesan__month=bln_b)
            days  = calendar.monthrange(thn_b, bln_b)[1]
            data_tgl_bln = []
            for day in range(1, days + 1):
                p    = p_bln.filter(waktu_pesan__day=day)
                jml  = p.count()
                pend = p.filter(status='selesai').aggregate(
                    t=Sum('total_harga'))['t'] or 0
                data_tgl_bln.append({
                    'hari': day,
                    'tgl_str': f'{thn_b}-{bln_b:02d}-{day:02d}',
                    'jumlah': jml, 'pendapatan': pend,
                })
                labels_t.append(str(day))
                val_jml_t.append(jml)
                val_pend_t.append(pend)
        except (ValueError, KeyError, IndexError):
            pass

    # Drill-down ke transaksi per tanggal
    pesanan_tgl, tgl_dipilih = None, None
    if ftgl:
        try:
            tgl_dipilih = datetime.date.fromisoformat(ftgl)
            pesanan_tgl = pesanan_tahun.filter(
                waktu_pesan__date=tgl_dipilih
            ).order_by('-waktu_pesan')
            thn_b = tgl_dipilih.year
            bln_b = tgl_dipilih.month
            bln_dipilih = {
                'str': f'{thn_b}-{bln_b:02d}',
                'nama': NAMA_BULAN[bln_b],
                'tahun': thn_b, 'bulan': bln_b,
            }
        except ValueError:
            pass

    return render(request, 'kasir/laporan_transaksi_tahun_ini.html', {
        'data_bulan'      : data_bulan,
        'data_tgl_bln'    : data_tgl_bln,
        'bln_dipilih'     : bln_dipilih,
        'pesanan_tgl'     : pesanan_tgl,
        'tgl_dipilih'     : tgl_dipilih,
        'fbulan'          : fbulan,
        'labels_b'        : json.dumps(labels_b),
        'val_jml_b'       : json.dumps(val_jml_b),
        'val_pend_b'      : json.dumps(val_pend_b),
        'labels_t'        : json.dumps(labels_t),
        'val_jml_t'       : json.dumps(val_jml_t),
        'val_pend_t'      : json.dumps(val_pend_t),
        'total_transaksi' : pesanan_tahun.count(),
        'total_pendapatan': pesanan_tahun.filter(status='selesai').aggregate(
            t=Sum('total_harga'))['t'] or 0,
        'judul': str(now.year),
    })


# ==============================
# LAPORAN TRANSAKSI — SEMUA
# ==============================
@login_required(login_url='/kasir/login/')
def lt_semua(request):
    ftahun = request.GET.get('tahun',   '')
    fbulan = request.GET.get('bulan',   '')
    ftgl   = request.GET.get('tanggal', '')

    semua = Pesanan.objects.select_related(
        'meja').prefetch_related('detail').all()

    # Level 1 — per tahun
    tahun_list = semua.dates('waktu_pesan', 'year')
    data_tahun, labels_y, val_jml_y, val_pend_y = [], [], [], []
    for t in tahun_list:
        p    = semua.filter(waktu_pesan__year=t.year)
        jml  = p.count()
        pend = p.filter(status='selesai').aggregate(
            t=Sum('total_harga'))['t'] or 0
        data_tahun.append({'tahun': t.year, 'jumlah': jml, 'pendapatan': pend})
        labels_y.append(str(t.year))
        val_jml_y.append(jml)
        val_pend_y.append(pend)

    # Level 2 — per bulan dalam tahun
    data_bulan, thn_dipilih = [], None
    labels_b, val_jml_b, val_pend_b = [], [], []
    if ftahun and not fbulan and not ftgl:
        try:
            thn_dipilih = int(ftahun)
            p_thn = semua.filter(waktu_pesan__year=thn_dipilih)
            for bln in range(1, 13):
                p    = p_thn.filter(waktu_pesan__month=bln)
                jml  = p.count()
                pend = p.filter(status='selesai').aggregate(
                    t=Sum('total_harga'))['t'] or 0
                if jml > 0:
                    data_bulan.append({
                        'bulan': bln, 'nama': NAMA_BULAN[bln],
                        'bln_str': f'{thn_dipilih}-{bln:02d}',
                        'jumlah': jml, 'pendapatan': pend,
                    })
                    labels_b.append(NAMA_BULAN[bln][:3])
                    val_jml_b.append(jml)
                    val_pend_b.append(pend)
        except ValueError:
            pass

    # Level 3 — per tanggal dalam bulan
    data_tgl, bln_dipilih = [], None
    labels_t, val_jml_t, val_pend_t = [], [], []
    if fbulan and not ftgl:
        try:
            thn_b, bln_b = int(fbulan.split('-')[0]), int(fbulan.split('-')[1])
            thn_dipilih  = thn_b
            bln_dipilih  = {'str': fbulan, 'nama': NAMA_BULAN[bln_b], 'tahun': thn_b, 'bulan': bln_b}
            p_bln = semua.filter(waktu_pesan__year=thn_b, waktu_pesan__month=bln_b)
            days  = calendar.monthrange(thn_b, bln_b)[1]
            for day in range(1, days + 1):
                p    = p_bln.filter(waktu_pesan__day=day)
                jml  = p.count()
                pend = p.filter(status='selesai').aggregate(
                    t=Sum('total_harga'))['t'] or 0
                data_tgl.append({
                    'hari': day,
                    'tgl_str': f'{thn_b}-{bln_b:02d}-{day:02d}',
                    'jumlah': jml, 'pendapatan': pend,
                })
                labels_t.append(str(day))
                val_jml_t.append(jml)
                val_pend_t.append(pend)
        except (ValueError, KeyError, IndexError):
            pass

    # Level 4 — transaksi per tanggal
    pesanan_tgl, tgl_dipilih = None, None
    if ftgl:
        try:
            tgl_dipilih = datetime.date.fromisoformat(ftgl)
            pesanan_tgl = semua.filter(
                waktu_pesan__date=tgl_dipilih
            ).order_by('-waktu_pesan')
            thn_dipilih = tgl_dipilih.year
            bln_dipilih = {
                'str': f'{tgl_dipilih.year}-{tgl_dipilih.month:02d}',
                'nama': NAMA_BULAN[tgl_dipilih.month],
                'tahun': tgl_dipilih.year, 'bulan': tgl_dipilih.month,
            }
        except ValueError:
            pass

    return render(request, 'kasir/laporan_transaksi_semua.html', {
        'data_tahun'  : data_tahun,
        'data_bulan'  : data_bulan,
        'data_tgl'    : data_tgl,
        'pesanan_tgl' : pesanan_tgl,
        'thn_dipilih' : thn_dipilih,
        'bln_dipilih' : bln_dipilih,
        'tgl_dipilih' : tgl_dipilih,
        'ftahun': ftahun, 'fbulan': fbulan, 'ftgl': ftgl,
        'labels_y' : json.dumps(labels_y),
        'val_jml_y': json.dumps(val_jml_y),
        'val_pend_y':json.dumps(val_pend_y),
        'labels_b' : json.dumps(labels_b),
        'val_jml_b': json.dumps(val_jml_b),
        'val_pend_b':json.dumps(val_pend_b),
        'labels_t' : json.dumps(labels_t),
        'val_jml_t': json.dumps(val_jml_t),
        'val_pend_t':json.dumps(val_pend_t),
        'total_transaksi' : semua.count(),
        'total_pendapatan': semua.filter(status='selesai').aggregate(
            t=Sum('total_harga'))['t'] or 0,
    })


# ==============================
# HELPER PENJUALAN
# ==============================
def _get_items(qs, kata_cari, urut):
    if kata_cari:
        qs = qs.filter(nama_menu_snapshot__icontains=kata_cari)
    items = qs.values('nama_menu_snapshot').annotate(
        total_terjual=Sum('jumlah'),
        total_omzet=Sum('subtotal'),
    )
    return items.order_by(
        '-total_omzet' if urut == 'omzet' else '-total_terjual'
    )


# ==============================
# LAPORAN PENJUALAN — HARI INI
# ==============================
@login_required(login_url='/kasir/login/')
def lp_hari_ini(request):
    today    = timezone.localtime(timezone.now()).date()
    kata     = request.GET.get('q', '').strip()
    urut     = request.GET.get('urut', 'terjual')

    detail = DetailPesanan.objects.filter(
        pesanan__waktu_pesan__date=today,
        pesanan__status='selesai',
    )
    items = _get_items(detail, kata, urut)

    return render(request, 'kasir/laporan_penjualan_hari_ini.html', {
        'items'     : items,
        'kata'      : kata,
        'urut'      : urut,
        'total_terjual' : detail.aggregate(t=Sum('jumlah'))['t'] or 0,
        'total_omzet'   : detail.aggregate(t=Sum('subtotal'))['t'] or 0,
        'judul' : f'Hari Ini — {_fmt_tgl(today)}',
    })


# ==============================
# LAPORAN PENJUALAN — BULAN INI
# ==============================
@login_required(login_url='/kasir/login/')
def lp_bulan_ini(request):
    now  = timezone.localtime(timezone.now())
    ftgl = request.GET.get('tanggal', '')
    kata = request.GET.get('q', '').strip()
    urut = request.GET.get('urut', 'terjual')

    detail_bulan = DetailPesanan.objects.filter(
        pesanan__waktu_pesan__year=now.year,
        pesanan__waktu_pesan__month=now.month,
        pesanan__status='selesai',
    )

    # Drill-down ke tanggal
    items_tgl, tgl_dipilih = None, None
    if ftgl:
        try:
            tgl_dipilih = datetime.date.fromisoformat(ftgl)
            d_tgl = detail_bulan.filter(pesanan__waktu_pesan__date=tgl_dipilih)
            items_tgl = _get_items(d_tgl, kata, urut)
        except ValueError:
            pass

    # List per tanggal
    days = calendar.monthrange(now.year, now.month)[1]
    data_tgl = []
    for day in range(1, days + 1):
        d = detail_bulan.filter(pesanan__waktu_pesan__day=day)
        data_tgl.append({
            'hari': day,
            'tgl_str': f'{now.year}-{now.month:02d}-{day:02d}',
            'jumlah': d.aggregate(t=Sum('jumlah'))['t'] or 0,
            'omzet' : d.aggregate(t=Sum('subtotal'))['t'] or 0,
        })

    return render(request, 'kasir/laporan_penjualan_bulan_ini.html', {
        'data_tgl'    : data_tgl,
        'items_tgl'   : items_tgl,
        'tgl_dipilih' : tgl_dipilih,
        'items_bulan' : _get_items(detail_bulan, kata, urut),
        'kata': kata, 'urut': urut,
        'total_terjual': detail_bulan.aggregate(t=Sum('jumlah'))['t'] or 0,
        'total_omzet'  : detail_bulan.aggregate(t=Sum('subtotal'))['t'] or 0,
        'judul': f'{NAMA_BULAN[now.month]} {now.year}',
    })


# ==============================
# LAPORAN PENJUALAN — TAHUN INI
# ==============================
@login_required(login_url='/kasir/login/')
def lp_tahun_ini(request):
    now    = timezone.localtime(timezone.now())
    fbulan = request.GET.get('bulan', '')
    ftgl   = request.GET.get('tanggal', '')
    kata   = request.GET.get('q', '').strip()
    urut   = request.GET.get('urut', 'terjual')

    detail_tahun = DetailPesanan.objects.filter(
        pesanan__waktu_pesan__year=now.year,
        pesanan__status='selesai',
    )

    # List per bulan
    data_bulan = []
    for bln in range(1, 13):
        d = detail_tahun.filter(pesanan__waktu_pesan__month=bln)
        data_bulan.append({
            'bulan': bln, 'nama': NAMA_BULAN[bln],
            'bln_str': f'{now.year}-{bln:02d}',
            'jumlah': d.aggregate(t=Sum('jumlah'))['t'] or 0,
            'omzet' : d.aggregate(t=Sum('subtotal'))['t'] or 0,
        })

    # Drill-down ke bulan
    data_tgl_bln, bln_dipilih, items_bln = None, None, None
    if fbulan and not ftgl:
        try:
            thn_b, bln_b = int(fbulan.split('-')[0]), int(fbulan.split('-')[1])
            bln_dipilih  = {'str': fbulan, 'nama': NAMA_BULAN[bln_b], 'tahun': thn_b, 'bulan': bln_b}
            d_bln = detail_tahun.filter(pesanan__waktu_pesan__month=bln_b)
            days  = calendar.monthrange(thn_b, bln_b)[1]
            data_tgl_bln = []
            for day in range(1, days + 1):
                d = d_bln.filter(pesanan__waktu_pesan__day=day)
                data_tgl_bln.append({
                    'hari': day,
                    'tgl_str': f'{thn_b}-{bln_b:02d}-{day:02d}',
                    'jumlah': d.aggregate(t=Sum('jumlah'))['t'] or 0,
                    'omzet' : d.aggregate(t=Sum('subtotal'))['t'] or 0,
                })
            items_bln = _get_items(d_bln, kata, urut)
        except (ValueError, KeyError, IndexError):
            pass

    # Drill-down ke tanggal
    items_tgl, tgl_dipilih = None, None
    if ftgl:
        try:
            tgl_dipilih = datetime.date.fromisoformat(ftgl)
            bln_dipilih = {
                'str': f'{tgl_dipilih.year}-{tgl_dipilih.month:02d}',
                'nama': NAMA_BULAN[tgl_dipilih.month],
                'tahun': tgl_dipilih.year, 'bulan': tgl_dipilih.month,
            }
            d_tgl    = detail_tahun.filter(pesanan__waktu_pesan__date=tgl_dipilih)
            items_tgl = _get_items(d_tgl, kata, urut)
        except ValueError:
            pass

    return render(request, 'kasir/laporan_penjualan_tahun_ini.html', {
        'data_bulan'  : data_bulan,
        'data_tgl_bln': data_tgl_bln,
        'bln_dipilih' : bln_dipilih,
        'items_bln'   : items_bln,
        'items_tgl'   : items_tgl,
        'tgl_dipilih' : tgl_dipilih,
        'fbulan': fbulan, 'ftgl': ftgl,
        'kata': kata, 'urut': urut,
        'total_terjual': detail_tahun.aggregate(t=Sum('jumlah'))['t'] or 0,
        'total_omzet'  : detail_tahun.aggregate(t=Sum('subtotal'))['t'] or 0,
        'judul': str(now.year),
    })


# ==============================
# LAPORAN PENJUALAN — SEMUA
# ==============================
@login_required(login_url='/kasir/login/')
def lp_semua(request):
    ftahun = request.GET.get('tahun',   '')
    fbulan = request.GET.get('bulan',   '')
    ftgl   = request.GET.get('tanggal', '')
    kata   = request.GET.get('q',       '').strip()
    urut   = request.GET.get('urut',    'terjual')

    semua_detail = DetailPesanan.objects.filter(pesanan__status='selesai')

    # Level 1 — per tahun
    tahun_list = Pesanan.objects.filter(
        status='selesai').dates('waktu_pesan', 'year')
    data_tahun = []
    for t in tahun_list:
        d = semua_detail.filter(pesanan__waktu_pesan__year=t.year)
        data_tahun.append({
            'tahun' : t.year,
            'jumlah': d.aggregate(t=Sum('jumlah'))['t'] or 0,
            'omzet' : d.aggregate(t=Sum('subtotal'))['t'] or 0,
        })

    # Level 2 — per bulan
    data_bulan, thn_dipilih = [], None
    if ftahun and not fbulan and not ftgl:
        try:
            thn_dipilih = int(ftahun)
            d_thn = semua_detail.filter(pesanan__waktu_pesan__year=thn_dipilih)
            for bln in range(1, 13):
                d = d_thn.filter(pesanan__waktu_pesan__month=bln)
                if d.exists():
                    data_bulan.append({
                        'bulan': bln, 'nama': NAMA_BULAN[bln],
                        'bln_str': f'{thn_dipilih}-{bln:02d}',
                        'jumlah': d.aggregate(t=Sum('jumlah'))['t'] or 0,
                        'omzet' : d.aggregate(t=Sum('subtotal'))['t'] or 0,
                    })
        except ValueError:
            pass

    # Level 3 — per tanggal
    data_tgl, bln_dipilih, items_bln = [], None, None
    if fbulan and not ftgl:
        try:
            thn_b, bln_b = int(fbulan.split('-')[0]), int(fbulan.split('-')[1])
            thn_dipilih  = thn_b
            bln_dipilih  = {'str': fbulan, 'nama': NAMA_BULAN[bln_b], 'tahun': thn_b, 'bulan': bln_b}
            d_bln = semua_detail.filter(
                pesanan__waktu_pesan__year=thn_b,
                pesanan__waktu_pesan__month=bln_b,
            )
            days = calendar.monthrange(thn_b, bln_b)[1]
            for day in range(1, days + 1):
                d = d_bln.filter(pesanan__waktu_pesan__day=day)
                data_tgl.append({
                    'hari': day,
                    'tgl_str': f'{thn_b}-{bln_b:02d}-{day:02d}',
                    'jumlah': d.aggregate(t=Sum('jumlah'))['t'] or 0,
                    'omzet' : d.aggregate(t=Sum('subtotal'))['t'] or 0,
                })
            items_bln = _get_items(d_bln, kata, urut)
        except (ValueError, KeyError, IndexError):
            pass

    # Level 4 — items per tanggal
    items_tgl, tgl_dipilih = None, None
    if ftgl:
        try:
            tgl_dipilih = datetime.date.fromisoformat(ftgl)
            thn_dipilih = tgl_dipilih.year
            bln_dipilih = {
                'str': f'{tgl_dipilih.year}-{tgl_dipilih.month:02d}',
                'nama': NAMA_BULAN[tgl_dipilih.month],
                'tahun': tgl_dipilih.year, 'bulan': tgl_dipilih.month,
            }
            d_tgl    = semua_detail.filter(pesanan__waktu_pesan__date=tgl_dipilih)
            items_tgl = _get_items(d_tgl, kata, urut)
        except ValueError:
            pass

    return render(request, 'kasir/laporan_penjualan_semua.html', {
        'data_tahun'  : data_tahun,
        'data_bulan'  : data_bulan,
        'data_tgl'    : data_tgl,
        'items_bln'   : items_bln,
        'items_tgl'   : items_tgl,
        'thn_dipilih' : thn_dipilih,
        'bln_dipilih' : bln_dipilih,
        'tgl_dipilih' : tgl_dipilih,
        'ftahun': ftahun, 'fbulan': fbulan, 'ftgl': ftgl,
        'kata': kata, 'urut': urut,
        'total_terjual': semua_detail.aggregate(t=Sum('jumlah'))['t'] or 0,
        'total_omzet'  : semua_detail.aggregate(t=Sum('subtotal'))['t'] or 0,
    })
@login_required(login_url='/kasir/login/')
def preview_qr(request, meja_id):
    """
    Tampilkan preview QR Code di halaman web.
    Return HTML dengan gambar QR.
    """
    try:
        meja = Meja.objects.get(id=meja_id)
    except Meja.DoesNotExist:
        return redirect('kasir:kelola_meja')

    # URL yang akan di-encode ke QR
    base_url  = request.build_absolute_uri('/')
    url_meja  = f"{base_url.rstrip('/')}"\
                f"/menu/{meja.nomor_meja}/"

    # Generate QR bytes
    qr_bytes  = generate_qr_dengan_label(
        url        = url_meja,
        nomor_meja = meja.nomor_meja,
        nama_kafe  = 'E-Order',
    )

    # Encode ke base64 untuk ditampilkan di HTML
    qr_base64 = base64.b64encode(qr_bytes).decode('utf-8')

    return render(request, 'kasir/preview_qr.html', {
        'meja'      : meja,
        'url_meja'  : url_meja,
        'qr_base64' : qr_base64,
    })


@login_required(login_url='/kasir/login/')
def download_qr(request, meja_id):
    """
    Download QR Code sebagai file PNG.
    """
    try:
        meja = Meja.objects.get(id=meja_id)
    except Meja.DoesNotExist:
        return redirect('kasir:kelola_meja')

    base_url = request.build_absolute_uri('/')
    url_meja = f"{base_url.rstrip('/')}/menu/{meja.nomor_meja}/"

    qr_bytes = generate_qr_dengan_label(
        url        = url_meja,
        nomor_meja = meja.nomor_meja,
        nama_kafe  = 'E-Order',
    )

    response = HttpResponse(qr_bytes, content_type='image/png')
    response['Content-Disposition'] = (
        f'attachment; filename="QR-Meja-{meja.nomor_meja}.png"'
    )
    return response

# ==============================
# PENGATURAN
# ==============================
@login_required(login_url='/kasir/login/')
def pengaturan(request):
    setting = Pengaturan.get()
    return render(request, 'kasir/pengaturan.html', {
        'setting'  : setting,
        'ip_server': get_client_ip(request),
    })


@login_required(login_url='/kasir/login/')
@require_POST
def simpan_pengaturan(request):
    setting = Pengaturan.get()

    # Info Kafe
    setting.nama_kafe   = request.POST.get('nama_kafe', '').strip()
    setting.alamat_kafe = request.POST.get('alamat_kafe', '').strip()
    setting.telp_kafe   = request.POST.get('telp_kafe', '').strip()
    setting.footer_struk= request.POST.get('footer_struk', '').strip()
    # Notifikasi Suara
    setting.notif_suara_pilihan = request.POST.get(
        'notif_suara_pilihan', '1'
    )
    volume_str = request.POST.get('notif_volume', '0.8')
    try:
        setting.notif_volume = float(volume_str)
    except ValueError:
        setting.notif_volume = 0.8
    setting.notif_loop = request.POST.get('notif_loop') == 'on'
    setting.save()
    # QRIS
    if 'gambar_qris' in request.FILES:
        setting.gambar_qris = request.FILES['gambar_qris']
    elif request.POST.get('hapus_qris') == '1':
        setting.gambar_qris = None

    # WiFi
    setting.wifi_aktif      = request.POST.get('wifi_aktif') == 'on'
    setting.wifi_ip_range   = request.POST.get('wifi_ip_range', '').strip()
    setting.pesan_wifi      = request.POST.get('pesan_wifi', '').strip()
    setting.wifi_password  = request.POST.get('wifi_password', '').strip()

    # Session
    durasi = request.POST.get('durasi_session', '10')
    setting.durasi_session = int(durasi) if durasi.isdigit() else 10

    import logging
    logging.warning(
        "[DEBUG SIMPAN PENGATURAN] durasi_post=%s durasi_session_set=%s",
        durasi,
        setting.durasi_session,
    )


    # Jam Operasional
    setting.cek_jam_operasional = (
        request.POST.get('cek_jam_operasional') == 'on'
    )
    jam_buka  = request.POST.get('jam_buka',  '08:00')
    jam_tutup = request.POST.get('jam_tutup', '22:00')
    if jam_buka:  setting.jam_buka  = jam_buka
    if jam_tutup: setting.jam_tutup = jam_tutup

    setting.save()

    import logging
    logging.warning(
        "[DEBUG SIMPAN PENGATURAN] durasi_session_after_save=%s",
        setting.durasi_session,
    )

    return redirect('kasir:pengaturan')



@login_required(login_url='/kasir/login/')
def get_ip_saya(request):
    """Return IP address client saat ini."""
    return JsonResponse({
        'ip': get_client_ip(request)
    })

@login_required(login_url='/kasir/login/')
def cetak_struk(request, nomor_pesanan):
    try:
        pesanan = Pesanan.objects.prefetch_related(
            'detail'
        ).select_related('meja').get(
            nomor_pesanan=nomor_pesanan
        )
    except Pesanan.DoesNotExist:
        return redirect('kasir:dashboard')

    # ✅ Ambil dari pengaturan
    setting = Pengaturan.get()

    return render(request, 'kasir/struk.html', {
        'pesanan'    : pesanan,
        'nama_kafe'  : setting.nama_kafe,
        'alamat_kafe': setting.alamat_kafe,
        'telp_kafe'  : setting.telp_kafe,
        'footer_kafe': setting.footer_struk,
    })
import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment,
    Border, Side
)
from openpyxl.utils import get_column_letter


def _style_header(cell):
    """Style untuk baris header Excel."""
    cell.font      = Font(bold=True, color='FFFFFF', size=11)
    cell.fill      = PatternFill(
        fill_type='solid', fgColor='1C1C1E'
    )
    cell.alignment = Alignment(
        horizontal='center', vertical='center'
    )
    cell.border    = Border(
        bottom=Side(style='thin', color='FFFFFF')
    )


def _style_subheader(cell):
    cell.font      = Font(bold=True, size=10)
    cell.fill      = PatternFill(
        fill_type='solid', fgColor='F5A623'
    )
    cell.alignment = Alignment(
        horizontal='center', vertical='center'
    )


def _auto_width(ws):
    """Auto-fit lebar kolom."""
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max_len + 4, 50)


# ==============================
# EXPORT TRANSAKSI
# ==============================
@login_required(login_url='/kasir/login/')
def export_transaksi_excel(request):
    """Export laporan transaksi ke Excel."""
    filter_status  = request.GET.get('status',  'semua')
    filter_tanggal = request.GET.get('tanggal', '')
    filter_bulan   = request.GET.get('bulan',   '')
    filter_tahun   = request.GET.get('tahun',   '')

    # Ambil data
    pesanan = Pesanan.objects.prefetch_related(
        'detail'
    ).select_related('meja').order_by('-waktu_pesan')

    if filter_status != 'semua':
        pesanan = pesanan.filter(status=filter_status)
    if filter_tanggal:
        try:
            tgl = datetime.date.fromisoformat(filter_tanggal)
            pesanan = pesanan.filter(waktu_pesan__date=tgl)
        except ValueError:
            pass
    if filter_bulan:
        try:
            thn, bln = filter_bulan.split('-')
            pesanan  = pesanan.filter(
                waktu_pesan__year=int(thn),
                waktu_pesan__month=int(bln)
            )
        except ValueError:
            pass
    if filter_tahun:
        try:
            pesanan = pesanan.filter(
                waktu_pesan__year=int(filter_tahun)
            )
        except ValueError:
            pass

    # Buat workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Laporan Transaksi'

    # Ambil info kafe
    try:
        setting   = Pengaturan.get()
        nama_kafe = setting.nama_kafe
    except Exception:
        nama_kafe = 'E-Order Café'

    # ===== JUDUL =====
    ws.merge_cells('A1:H1')
    judul        = ws['A1']
    judul.value  = f'LAPORAN TRANSAKSI — {nama_kafe}'
    judul.font   = Font(bold=True, size=14)
    judul.alignment = Alignment(horizontal='center')

    ws.merge_cells('A2:H2')
    sub        = ws['A2']
    sub.value  = (
        f'Diekspor: '
        f'{timezone.localtime(timezone.now()).strftime("%d/%m/%Y %H:%M")}'
    )
    sub.font      = Font(size=10, color='6B7280')
    sub.alignment = Alignment(horizontal='center')

    ws.append([])  # baris kosong

    # ===== RINGKASAN =====
    ws.merge_cells('A4:H4')
    ws['A4'].value     = 'RINGKASAN'
    _style_subheader(ws['A4'])

    total_semua   = pesanan.count()
    total_selesai = pesanan.filter(status='selesai').count()
    total_pend    = pesanan.filter(
        status='selesai'
    ).aggregate(t=Sum('total_harga'))['t'] or 0

    ws.append(['Total Transaksi', total_semua,
               '', 'Total Selesai', total_selesai,
               '', 'Total Pendapatan',
               f'Rp {total_pend:,.0f}'])
    ws.append([])

    # ===== HEADER TABEL =====
    headers = [
        'No', 'Nomor Pesanan', 'Meja', 'Waktu',
        'Status', 'Metode Bayar', 'Item',
        'Total (Rp)'
    ]
    ws.append(headers)
    for i, _ in enumerate(headers, 1):
        _style_header(ws.cell(row=ws.max_row, column=i))

    ws.row_dimensions[ws.max_row].height = 25

    # ===== DATA =====
    STATUS_LABEL = {
        'pending' : 'Pending',
        'diproses': 'Diproses',
        'selesai' : 'Selesai',
        'dibatal' : 'Dibatalkan',
    }
    METODE_LABEL = {
        'tunai': 'Tunai',
        'qris' : 'QRIS',
    }

    for no, p in enumerate(pesanan, 1):
        item_str = ' | '.join([
            f"{d.jumlah}x {d.nama_menu_snapshot}"
            for d in p.detail.all()
        ])
        row = [
            no,
            p.nomor_pesanan,
            f"Meja {p.meja.nomor_meja}" if p.meja else '-',
            timezone.localtime(
                p.waktu_pesan
            ).strftime('%d/%m/%Y %H:%M'),
            STATUS_LABEL.get(p.status, p.status),
            METODE_LABEL.get(p.metode_bayar, p.metode_bayar),
            item_str,
            p.total_harga,
        ]
        ws.append(row)

        # Warna baris berdasarkan status
        baris = ws.max_row
        warna = {
            'pending' : 'FEF9C3',
            'diproses': 'DBEAFE',
            'selesai' : 'DCFCE7',
            'dibatal' : 'FEE2E2',
        }.get(p.status, 'FFFFFF')

        for col in range(1, 9):
            cell      = ws.cell(row=baris, column=col)
            cell.fill = PatternFill(
                fill_type='solid', fgColor=warna
            )
            cell.alignment = Alignment(vertical='center')

    # Format kolom total sebagai angka
    for row in ws.iter_rows(
        min_row=8, max_row=ws.max_row,
        min_col=8, max_col=8
    ):
        for cell in row:
            if isinstance(cell.value, (int, float)):
                cell.number_format = '#,##0'

    _auto_width(ws)

    # ===== SHEET DETAIL =====
    ws2        = wb.create_sheet('Detail Item')
    ws2.title  = 'Detail Item'

    headers2 = [
        'No', 'Nomor Pesanan', 'Meja', 'Waktu',
        'Nama Menu', 'Jumlah', 'Harga Satuan (Rp)',
        'Subtotal (Rp)'
    ]
    ws2.append(headers2)
    for i, _ in enumerate(headers2, 1):
        _style_header(ws2.cell(row=1, column=i))

    no = 1
    for p in pesanan:
        for item in p.detail.all():
            ws2.append([
                no,
                p.nomor_pesanan,
                f"Meja {p.meja.nomor_meja}" if p.meja else '-',
                timezone.localtime(
                    p.waktu_pesan
                ).strftime('%d/%m/%Y %H:%M'),
                item.nama_menu_snapshot,
                item.jumlah,
                item.harga_snapshot,
                item.subtotal,
            ])
            no += 1

    for row in ws2.iter_rows(
        min_row=2, max_row=ws2.max_row,
        min_col=7, max_col=8
    ):
        for cell in row:
            cell.number_format = '#,##0'

    _auto_width(ws2)

    # Response
    response = HttpResponse(
        content_type=(
            'application/vnd.openxmlformats-'
            'officedocument.spreadsheetml.sheet'
        )
    )
    nama_file = (
        f'Laporan-Transaksi-'
        f'{timezone.localtime(timezone.now()).strftime("%Y%m%d-%H%M")}'
        f'.xlsx'
    )
    response['Content-Disposition'] = (
        f'attachment; filename="{nama_file}"'
    )
    wb.save(response)
    return response


# ==============================
# EXPORT PENJUALAN
# ==============================
@login_required(login_url='/kasir/login/')
def export_penjualan_excel(request):
    """Export laporan penjualan per item ke Excel."""
    filter_bulan = request.GET.get('bulan', '')
    filter_tahun = request.GET.get('tahun', '')

    detail = DetailPesanan.objects.filter(
        pesanan__status='selesai'
    ).select_related('pesanan', 'pesanan__meja')

    if filter_bulan:
        try:
            thn, bln = filter_bulan.split('-')
            detail   = detail.filter(
                pesanan__waktu_pesan__year=int(thn),
                pesanan__waktu_pesan__month=int(bln)
            )
        except ValueError:
            pass

    if filter_tahun:
        try:
            detail = detail.filter(
                pesanan__waktu_pesan__year=int(filter_tahun)
            )
        except ValueError:
            pass

    # Buat workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Laporan Penjualan'

    try:
        nama_kafe = Pengaturan.get().nama_kafe
    except Exception:
        nama_kafe = 'E-Order Café'

    # Judul
    ws.merge_cells('A1:F1')
    ws['A1'].value     = f'LAPORAN PENJUALAN — {nama_kafe}'
    ws['A1'].font      = Font(bold=True, size=14)
    ws['A1'].alignment = Alignment(horizontal='center')

    ws.merge_cells('A2:F2')
    ws['A2'].value = (
        f'Diekspor: '
        f'{timezone.localtime(timezone.now()).strftime("%d/%m/%Y %H:%M")}'
    )
    ws['A2'].font      = Font(size=10, color='6B7280')
    ws['A2'].alignment = Alignment(horizontal='center')

    ws.append([])

    # Ringkasan per menu
    ws.merge_cells('A4:F4')
    ws['A4'].value = 'RINGKASAN PER MENU'
    _style_subheader(ws['A4'])

    headers = [
        'No', 'Nama Menu',
        'Total Terjual', 'Harga Rata-rata (Rp)',
        'Total Omzet (Rp)', 'Persentase'
    ]
    ws.append(headers)
    for i, _ in enumerate(headers, 1):
        _style_header(ws.cell(row=ws.max_row, column=i))

    # Group by menu
    from django.db.models import Avg
    items = detail.values('nama_menu_snapshot').annotate(
        total_terjual = Sum('jumlah'),
        total_omzet   = Sum('subtotal'),
        rata_harga    = Avg('harga_snapshot'),
    ).order_by('-total_terjual')

    total_omzet_semua = sum(
        i['total_omzet'] for i in items
    ) or 1

    for no, item in enumerate(items, 1):
        persen = (item['total_omzet'] / total_omzet_semua) * 100
        ws.append([
            no,
            item['nama_menu_snapshot'],
            item['total_terjual'],
            round(item['rata_harga'], 0),
            item['total_omzet'],
            f"{persen:.1f}%",
        ])

        baris = ws.max_row
        if no % 2 == 0:
            for col in range(1, 7):
                ws.cell(row=baris, column=col).fill = PatternFill(
                    fill_type='solid', fgColor='F9FAFB'
                )

    # Format angka
    for row in ws.iter_rows(
        min_row=6, max_row=ws.max_row,
        min_col=4, max_col=5
    ):
        for cell in row:
            if isinstance(cell.value, (int, float)):
                cell.number_format = '#,##0'

    ws.append([])

    # Total
    total_row = ws.max_row + 1
    ws.cell(row=total_row, column=1).value = 'TOTAL'
    ws.cell(row=total_row, column=1).font  = Font(bold=True)
    ws.cell(row=total_row, column=3).value = (
        detail.aggregate(t=Sum('jumlah'))['t'] or 0
    )
    ws.cell(row=total_row, column=3).font  = Font(bold=True)
    ws.cell(row=total_row, column=5).value = (
        detail.aggregate(t=Sum('subtotal'))['t'] or 0
    )
    ws.cell(row=total_row, column=5).font         = Font(bold=True)
    ws.cell(row=total_row, column=5).number_format = '#,##0'

    # Sheet 2: Detail transaksi per item
    ws2       = wb.create_sheet('Detail Transaksi')
    headers2  = [
        'No', 'Tanggal', 'Nomor Pesanan',
        'Meja', 'Nama Menu', 'Jumlah',
        'Harga (Rp)', 'Subtotal (Rp)'
    ]
    ws2.append(headers2)
    for i, _ in enumerate(headers2, 1):
        _style_header(ws2.cell(row=1, column=i))

    for no, d in enumerate(
        detail.order_by('-pesanan__waktu_pesan'), 1
    ):
        ws2.append([
            no,
            timezone.localtime(
                d.pesanan.waktu_pesan
            ).strftime('%d/%m/%Y'),
            d.pesanan.nomor_pesanan,
            (f"Meja {d.pesanan.meja.nomor_meja}"
             if d.pesanan.meja else '-'),
            d.nama_menu_snapshot,
            d.jumlah,
            d.harga_snapshot,
            d.subtotal,
        ])

    for row in ws2.iter_rows(
        min_row=2, max_row=ws2.max_row,
        min_col=7, max_col=8
    ):
        for cell in row:
            cell.number_format = '#,##0'

    _auto_width(ws)
    _auto_width(ws2)

    response = HttpResponse(
        content_type=(
            'application/vnd.openxmlformats-'
            'officedocument.spreadsheetml.sheet'
        )
    )
    nama_file = (
        f'Laporan-Penjualan-'
        f'{timezone.localtime(timezone.now()).strftime("%Y%m%d-%H%M")}'
        f'.xlsx'
    )
    response['Content-Disposition'] = (
        f'attachment; filename="{nama_file}"'
    )
    wb.save(response)
    return response